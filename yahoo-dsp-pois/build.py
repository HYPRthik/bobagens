#!/usr/bin/env python3
"""
Gera os arquivos de geofencing da Yahoo DSP a partir da planilha de POIs unificados.

Contexto: o upload original falhou porque o arquivo foi enviado no campo de
ENDERECO da DSP (geofencing address list). A DSP tentou geocodificar as colunas
como endereco em vez de ler as coordenadas. Este script produz o formato que o
campo de LATITUDE/LONGITUDE espera: 3 colunas, sem cabecalho, ASCII, CRLF.

Uso:  python3 build.py [caminho/do/csv/de/origem]
"""
import collections
import csv
import io
import os
import re
import shutil
import sys
import unicodedata
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, "origem_pois_unificados.csv")
XLSX = os.path.join(BASE, "HYPR_JLR_POIs_Unificados_LIMPO.xlsx")
DECIMALS = 6          # ~11 cm; mata o ruido de float do Excel/Sheets
BR_BBOX = (-33.8, 5.3, -74.0, -28.8)   # lat_min, lat_max, lon_min, lon_max


def num(x, d=DECIMALS):
    v = round(float(x), d)
    return float(f"{v:.{d}f}".rstrip("0").rstrip("."))


def fmt(v, d=DECIMALS):
    return f"{v:.{d}f}".rstrip("0").rstrip(".") or "0"


def slug(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")


def csvlines(path):
    """Conta registros reais terminados em CRLF (modo texto normalizaria CRLF->LF)."""
    return len([l for l in open(path, "rb").read().split(b"\r\n") if l.strip()])


def write_geofence(path, records):
    """Arquivo de geofence lat/long: 3 colunas, SEM cabecalho, ASCII, CRLF.
    Deduplica por coordenada — a DSP nao aceita o mesmo geofence duas vezes."""
    seen, out = set(), []
    for lat, lon, rad, _est, _poi in records:
        if (lat, lon) in seen:
            continue
        seen.add((lat, lon))
        out.append((fmt(lat), fmt(lon), fmt(rad, 2)))
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="ascii") as f:
        csv.writer(f, lineterminator="\r\n").writerows(out)
    return len(out)


# ---------------------------------------------------------------- carga + limpeza
raw = list(csv.reader(io.StringIO(open(SRC, encoding="utf-8-sig").read())))
header, body = raw[0], raw[1:]
assert [h.strip() for h in header] == ["Latitude", "Longitude", "Radius Distance", "Estado", "POI"], header

records, seen, rejected = [], set(), []
for i, r in enumerate(body, start=2):
    try:
        lat, lon, rad = num(r[0]), num(r[1]), num(r[2], 2)
    except (ValueError, IndexError):
        rejected.append((i, r, "coordenada ilegivel"))
        continue
    if not (BR_BBOX[0] <= lat <= BR_BBOX[1] and BR_BBOX[2] <= lon <= BR_BBOX[3]):
        rejected.append((i, r, "fora do Brasil"))
        continue
    est, poi = r[3].strip(), r[4].strip()
    key = (lat, lon, poi)          # 1 linha por geofence POR CATEGORIA
    if key in seen:
        continue
    seen.add(key)
    records.append((lat, lon, rad, est, poi))

# uma coordenada nao pode pertencer a dois Estados
por_coord = collections.defaultdict(set)
for lat, lon, _r, est, _p in records:
    por_coord[(lat, lon)].add(est)
assert not [k for k, v in por_coord.items() if len(v) > 1], "coordenada com Estado conflitante"

# flag de primeira ocorrencia da coordenada (espelha a coluna F da planilha)
vistos, primeira = set(), []
for lat, lon, _r, _e, _p in records:
    novo = (lat, lon) not in vistos
    vistos.add((lat, lon))
    primeira.append(1 if novo else 0)

ESTADOS = sorted({r[3] for r in records})
POIS = sorted({r[4] for r in records})

# ---------------------------------------------------------------- arquivos de upload
n_todos = write_geofence(os.path.join(BASE, "upload", "pois_unificados_TODOS.csv"), records)
n_est = {e: write_geofence(os.path.join(BASE, "por_estado", f"{slug(e)}.csv"),
                           [r for r in records if r[3] == e]) for e in ESTADOS}
n_poi = {p: write_geofence(os.path.join(BASE, "por_poi", f"{slug(p)}.csv"),
                           [r for r in records if r[4] == p]) for p in POIS}

# ---------------------------------------------------------------- planilha de referencia
A = lambda **k: Font(name="Arial", **k)
HF = PatternFill("solid", fgColor="1F3864")
HFONT = A(bold=True, color="FFFFFF", size=10)
T = Side(style="thin", color="D9D9D9")
BORD = Border(left=T, right=T, top=T, bottom=T)
AZUL = A(bold=True, size=10, color="1F3864")

wb = Workbook()
ws = wb.active
ws.title = "LEIA-ME"
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 98
LEIA = [
 ("HYPR / JLR — POIs Unificados · versão corrigida para Yahoo DSP", ""), ("", ""),
 ("PROBLEMA", "O arquivo foi enviado no campo de upload por ENDEREÇO (geofencing address list)."),
 ("", "A DSP tentou geocodificar as colunas como endereço, em vez de ler as coordenadas."),
 ("", "Prova 1: a própria linha de CABEÇALHO ('Latitude,Longitude,...') voltou como 'successful'"),
 ("", "— a DSP geocodificou o texto do cabeçalho como se fosse um endereço."),
 ("", "Prova 2: o motivo do erro acompanha exatamente a coluna Estado:"),
 ("", "        Estado = 'SP' (UF válida)  ->  'Incomplete address'            1.030 linhas"),
 ("", "        Estado = 'Demais Praças'   ->  'Unknown error'                 1.021 linhas"),
 ("", "        Estado = 'Demais Praças'   ->  'Cannot match full address'       119 linhas"),
 ("", "Os 3 'successful' foram coincidência de geocodificação, não acerto de coordenada."),
 ("", "Outras 513 linhas foram descartadas em silêncio, sem nem aparecer no relatório."), ("", ""),
 ("SOLUÇÃO", "Subir os arquivos da pasta /upload no campo de LATITUDE/LONGITUDE da DSP."),
 ("", "Formato: 3 colunas (lat, lon, raio), SEM cabeçalho, sem acento, quebra de linha CRLF."), ("", ""),
 ("LIMPEZA APLICADA", "1. Cabeçalho removido — a DSP lê a 1a linha como dado."),
 ("", "2. Colunas 'Estado' e 'POI' retiradas do upload — viravam campos de endereço."),
 ("", f"3. Duplicatas removidas: {len(body):,} linhas originais -> {n_todos:,} geofences únicos."),
 ("", f"4. Coordenadas arredondadas para {DECIMALS} casas (~11 cm), eliminando ruído de float"),
 ("", "        ex.: -46,72234770000001  ->  -46,722348"),
 ("", "5. Validação: 100% das coordenadas são válidas e caem dentro do Brasil."), ("", ""),
 ("ATENÇÃO — RAIO", "O valor 0,3 foi mantido como estava. Confirme a UNIDADE no painel da DSP:"),
 ("", "se a DSP interpretar em MILHAS, 0,3 = 483 m em vez dos 300 m pretendidos."), ("", ""),
 (f"{len(records):,} x {n_todos:,}".replace(",", "."),
  f"A aba 'POIs_Limpos' tem {len(records):,} linhas e o arquivo único de upload tem {n_todos:,}.".replace(",", ".")),
 ("", f"A diferença são {len(records)-n_todos} coordenadas que pertencem a DUAS categorias (ex.: um clube"),
 ("", "que também é golf club). Elas aparecem nos dois arquivos por POI — correto, são listas"),
 ("", "de targeting separadas — mas só uma vez no arquivo único, que não aceita geofence repetido."), ("", ""),
 ("NESTE ARQUIVO", "'POIs_Limpos' = base completa com Estado e POI (referência, não é o arquivo de upload)."),
 ("", "'Resumo' = geofences por Estado e por categoria de POI."),
 ("", "Arquivos de upload prontos: /upload, /por_estado e /por_poi (CSV)."),
]
for i, (a, b) in enumerate(LEIA, 1):
    ws.cell(i, 1, a).font = A(bold=True, size=10)
    c = ws.cell(i, 2, b)
    c.font = A(size=10)
    c.alignment = Alignment(vertical="top")
ws["A1"].font = A(bold=True, size=13, color="1F3864")
for r in (3, 14, 17, 24, 27, 32):
    ws.cell(r, 1).font = AZUL
for r in (24, 25):
    ws.cell(r, 2).font = A(size=10, bold=True, color="C00000")

# --- aba POIs_Limpos ---
w2 = wb.create_sheet("POIs_Limpos")
COLS = [("Latitude", 14), ("Longitude", 14), ("Radius Distance", 17),
        ("Estado", 17), ("POI", 20), ("Coordenada única", 17)]
for c, (h, wd) in enumerate(COLS, 1):
    cell = w2.cell(1, c, h)
    cell.font = HFONT
    cell.fill = HF
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    w2.column_dimensions[get_column_letter(c)].width = wd
for rec in records:
    w2.append(list(rec))
for r in range(2, len(records) + 2):
    # 1 na primeira vez que a coordenada aparece, 0 nas repetições -> permite contar únicos
    w2.cell(r, 6, f"=IF(COUNTIFS($A$2:$A{r},$A{r},$B$2:$B{r},$B{r})=1,1,0)")
for row in w2.iter_rows(min_row=2):
    for c in row:
        c.font = A(size=10)
        c.border = BORD
        if c.column <= 2:
            c.number_format = "0.000000"
        elif c.column == 3:
            c.number_format = "0.0"
        elif c.column == 6:
            c.number_format = "0"
            c.alignment = Alignment(horizontal="center")
N = w2.max_row
w2.freeze_panes = "A2"
w2.auto_filter.ref = f"A1:F{N}"

# --- aba Resumo ---
w3 = wb.create_sheet("Resumo")
w3.column_dimensions["A"].width = 30
for col in "BCD":
    w3.column_dimensions[col].width = 16
w3.merge_cells("A1:D1")
w3["A1"] = "Geofences por categoria de POI"
w3["A1"].font = A(bold=True, size=12, color="FFFFFF")
w3["A1"].fill = HF
w3["A1"].alignment = Alignment(horizontal="left", vertical="center")
for c, h in enumerate(["POI", "SP", "Demais Praças", "Total"], 1):
    cell = w3.cell(2, c, h)
    cell.font = HFONT
    cell.fill = HF
    cell.alignment = Alignment(horizontal="center")
for i, p in enumerate(POIS):
    r = 3 + i
    w3.cell(r, 1, p).font = A(size=10)
    w3.cell(r, 2, f"=COUNTIFS(POIs_Limpos!$E$2:$E${N},$A{r},POIs_Limpos!$D$2:$D${N},B$2)")
    w3.cell(r, 3, f"=COUNTIFS(POIs_Limpos!$E$2:$E${N},$A{r},POIs_Limpos!$D$2:$D${N},C$2)")
    w3.cell(r, 4, f"=SUM(B{r}:C{r})")
    for c in range(2, 5):
        w3.cell(r, c).font = A(size=10)
        w3.cell(r, c).number_format = "#,##0"
    for c in range(1, 5):
        w3.cell(r, c).border = BORD
TR = 3 + len(POIS)
w3.cell(TR, 1, "TOTAL (POI x coordenada)").font = A(bold=True, size=10)
UR = TR + 1
w3.cell(UR, 1, "Geofences ÚNICOS (upload)").font = A(bold=True, size=10)
for c in range(2, 5):
    L = get_column_letter(c)
    t = w3.cell(TR, c, f"=SUM({L}3:{L}{TR-1})")
    if c < 4:
        u = w3.cell(UR, c, f'=SUMIFS(POIs_Limpos!$F$2:$F${N},POIs_Limpos!$D$2:$D${N},{L}$2)')
    else:
        u = w3.cell(UR, c, f"=SUM(B{UR}:C{UR})")
    for cell, fill in ((t, "DEEAF6"), (u, "FFF2CC")):
        cell.font = A(bold=True, size=10)
        cell.number_format = "#,##0"
        cell.fill = PatternFill("solid", fgColor=fill)
w3.cell(TR, 1).fill = PatternFill("solid", fgColor="DEEAF6")
w3.cell(UR, 1).fill = PatternFill("solid", fgColor="FFF2CC")
for r in (TR, UR):
    for c in range(1, 5):
        w3.cell(r, c).border = BORD
notas = [
    f"TOTAL = {len(records):,} pares POI x coordenada — soma dos arquivos da pasta /por_poi.".replace(",", "."),
    f"ÚNICOS = {n_todos:,} coordenadas distintas — é o tamanho de upload/pois_unificados_TODOS.csv.".replace(",", "."),
    f"A diferença de {len(records)-n_todos} são coordenadas que pertencem a duas categorias. Ver aba LEIA-ME.",
    "Coluna F de 'POIs_Limpos' marca com 1 a primeira ocorrência de cada coordenada.",
]
for i, t in enumerate(notas):
    w3.cell(UR + 2 + i, 1, t).font = A(size=9, italic=True)
w3.freeze_panes = "A3"
wb.save(XLSX)

# ---------------------------------------------------------------- valores em cache
# O LibreOffice nao carrega arquivo nenhum neste ambiente, entao os resultados sao
# calculados aqui e gravados como <v> ao lado de cada <f>. As formulas continuam
# vivas (calcPr fullCalcOnLoad=1 faz o Excel recalcular ao abrir); o cache serve
# para pandas/visualizadores que so leem valor.
truth = collections.Counter()
tot_par = collections.Counter()
tot_uni = collections.Counter()
for (lat, lon, _r, est, poi), first in zip(records, primeira):
    truth[(poi, est)] += 1
    tot_par[est] += 1
    tot_uni[est] += first

vals = {}
for i, p in enumerate(POIS):
    r = 3 + i
    sp, dp = truth[(p, "SP")], truth[(p, "Demais Praças")]
    vals[f"B{r}"], vals[f"C{r}"], vals[f"D{r}"] = sp, dp, sp + dp
vals[f"B{TR}"], vals[f"C{TR}"] = tot_par["SP"], tot_par["Demais Praças"]
vals[f"D{TR}"] = sum(tot_par.values())
vals[f"B{UR}"], vals[f"C{UR}"] = tot_uni["SP"], tot_uni["Demais Praças"]
vals[f"D{UR}"] = sum(tot_uni.values())
col_f = {f"F{r}": primeira[r - 2] for r in range(2, N + 1)}

PAT = re.compile(r'(<c r="([A-Z]+\d+)"[^>]*>\s*<f>.*?</f>\s*)<v\s*/>(\s*</c>)', re.S)
shutil.copy(XLSX, XLSX + ".bak")
zin = zipfile.ZipFile(XLSX + ".bak")
zout = zipfile.ZipFile(XLSX, "w", zipfile.ZIP_DEFLATED)
sheet_of = {"POIs_Limpos": "xl/worksheets/sheet2.xml", "Resumo": "xl/worksheets/sheet3.xml"}
counts = collections.Counter()
for item in zin.infolist():
    buf = zin.read(item.filename)
    table = vals if item.filename == sheet_of["Resumo"] else (
        col_f if item.filename == sheet_of["POIs_Limpos"] else None)
    if table:
        def fix(m, table=table, key=item.filename):
            ref = m.group(2)
            if ref in table:
                counts[key] += 1
                return f"{m.group(1)}<v>{table[ref]}</v>{m.group(3)}"
            return m.group(0)
        buf = PAT.sub(fix, buf.decode()).encode()
    zout.writestr(item, buf)
zout.close()
zin.close()
os.remove(XLSX + ".bak")
assert counts[sheet_of["Resumo"]] == len(vals), counts
assert counts[sheet_of["POIs_Limpos"]] == len(col_f), counts

# ---------------------------------------------------------------- validacao final
falhas = []
for path in sorted([os.path.join(dp, f) for dp, _, fs in os.walk(BASE)
                    for f in fs if f.endswith(".csv") and "origem" not in f]):
    b = open(path, "rb").read()
    rows_ = list(csv.reader(io.StringIO(b.decode("ascii"))))
    rel = os.path.relpath(path, BASE)
    if any(len(r) != 3 for r in rows_):
        falhas.append(f"{rel}: coluna != 3")
    if rows_ and rows_[0][0].lower().startswith("lat"):
        falhas.append(f"{rel}: TEM CABECALHO")
    if b.count(b"\r\n") != len(rows_):
        falhas.append(f"{rel}: quebra de linha != CRLF")
    if len({(r[0], r[1]) for r in rows_}) != len(rows_):
        falhas.append(f"{rel}: duplicatas")
    for r in rows_:
        if not (BR_BBOX[0] <= float(r[0]) <= BR_BBOX[1] and BR_BBOX[2] <= float(r[1]) <= BR_BBOX[3]):
            falhas.append(f"{rel}: coordenada fora do Brasil")
            break

wv = load_workbook(XLSX, data_only=True)["Resumo"]
wf = load_workbook(XLSX)["Resumo"]
if not str(wf.cell(3, 2).value).startswith("=COUNTIFS"):
    falhas.append("planilha: formula perdida")
for i, p in enumerate(POIS):
    if wv.cell(3 + i, 4).value != n_poi[p]:
        falhas.append(f"planilha: Resumo {p} != por_poi/{slug(p)}.csv")
if wv.cell(UR, 4).value != n_todos:
    falhas.append("planilha: unicos != arquivo de upload")
for e in ESTADOS:
    col = 2 if e == "SP" else 3
    if wv.cell(UR, col).value != n_est[e]:
        falhas.append(f"planilha: unicos {e} != por_estado/{slug(e)}.csv")
if sum(n_est.values()) != n_todos:
    falhas.append("soma por_estado != arquivo unico")
if sum(n_poi.values()) != len(records):
    falhas.append("soma por_poi != linhas da planilha")

print(f"origem            : {len(body):,} linhas".replace(",", "."))
print(f"descartadas       : {len(rejected)}")
print(f"planilha          : {len(records):,} pares POI x coordenada".replace(",", "."))
print(f"upload (único)    : {n_todos:,} geofences".replace(",", "."))
print(f"por estado        : " + "  ".join(f"{e}={v}" for e, v in n_est.items()))
print(f"por POI           : {len(n_poi)} arquivos, {sum(n_poi.values()):,} linhas".replace(",", "."))
print("\nvalidação         : " + ("TUDO OK" if not falhas else "FALHAS:\n  " + "\n  ".join(falhas)))
sys.exit(1 if falhas else 0)
